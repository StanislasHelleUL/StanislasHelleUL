# -*- coding: utf-8 -*-
"""
Created on Tue Feb  8 13:04:21 2022

@author: stan
"""

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtPrintSupport import QPrinter
from urllib.error import URLError, HTTPError
from urllib.request import urlopen as url
import subprocess as sp
import openpyxl
import qdarktheme

import UniProtWindow
import secondWindow_v2
import PlotWindow
import choose_MW
import InSilicoDigestion
from DictioProt import aminoDict
from Protein_functions import peptideMW, mzPeakMassif


class Ui_MainWindow(object):
    
    def setupUi(self, MainWindow):
        
        self.settings = QtCore.QSettings("LabGUI_app_v1.4", "app_settings")
        
        MainWindow.setObjectName("MainWindow")
        
        try:self.lightMode = True if self.settings.value('lightMode') == 'true' else False       
        except: self.lightMode = True
        
        MainWindow.resize(800, 600) 
        
        self.mainFont = QtGui.QFont()
        self.mainFont.setFamily("Century Gothic")
        self.mainFont.setPointSize(11)
        
        self.fastaFont = QtGui.QFont()
        self.fastaFont.setFamily("Consolas")
        self.fastaFont.setPointSize(11)
        
        self.LayoutWidgetGeometry = QtCore.QRect(9, 9, 761, 481)
        
        self.mainIcon = QtGui.QIcon()
        self.mainIcon.addPixmap(QtGui.QPixmap("data/SSPC_brand-01.png"))
        MainWindow.setWindowIcon(self.mainIcon)        
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")

        self.tabWidget = QtWidgets.QTabWidget(self.centralwidget)
        self.tabWidget.setGeometry(QtCore.QRect(10, 10, 781, 541))
        self.tabWidget.setObjectName("tabWidget")
        
        self.tab_FindWeight = QtWidgets.QWidget()
        self.tab_FindWeight.setObjectName("tab_FindWeight")
        self.verticalLayoutWidget = QtWidgets.QWidget(self.tab_FindWeight)
        self.verticalLayoutWidget.setGeometry(self.LayoutWidgetGeometry)
        self.verticalLayoutWidget.setObjectName("verticalLayoutWidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.verticalLayoutWidget)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout.setObjectName("verticalLayout")
        self.label = QtWidgets.QLabel(self.verticalLayoutWidget)
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName("label")
        self.verticalLayout.addWidget(self.label)
        self.spinBox = QtWidgets.QSpinBox(self.verticalLayoutWidget)
        self.spinBox.setAlignment(QtCore.Qt.AlignCenter)
        self.spinBox.setMinimum(1)
        self.spinBox.setMaximum(99)
        self.spinBox.setObjectName("spinBox")
        self.spinBox.setFont(self.mainFont)
        self.verticalLayout.addWidget(self.spinBox)
        self.pushButton = QtWidgets.QPushButton(self.verticalLayoutWidget)
        self.pushButton.setObjectName("pushButton")
        self.verticalLayout.addWidget(self.pushButton)
        self.tabWidget.addTab(self.tab_FindWeight, "")
        
        self.tab_FindProt = QtWidgets.QWidget()
        self.tab_FindProt.setObjectName("tab_FindProt")
        self.verticalLayoutWidget2 = QtWidgets.QWidget(self.tab_FindProt)
        self.verticalLayoutWidget2.setGeometry(self.LayoutWidgetGeometry)
        self.verticalLayoutWidget2.setObjectName("verticalLayoutWidget2")
        self.verticalLayout2 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget2)
        self.verticalLayout2.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout2.setObjectName("verticalLayout2")
        self.label2 = QtWidgets.QLabel(self.verticalLayoutWidget2)
        self.label2.setAlignment(QtCore.Qt.AlignCenter)
        self.label2.setObjectName("label2")
        self.verticalLayout2.addWidget(self.label2)
        self.prot_lineEdit = QtWidgets.QLineEdit(self.verticalLayoutWidget2)
        self.prot_lineEdit.setObjectName("prot_lineEdit")
        self.verticalLayout2.addWidget(self.prot_lineEdit)
        self.text_rb = QtWidgets.QRadioButton(self.verticalLayoutWidget2)
        self.text_rb.setObjectName("text_rb")
        self.verticalLayout2.addWidget(self.text_rb)
        self.fasta_rb = QtWidgets.QRadioButton(self.verticalLayoutWidget2)
        self.fasta_rb.setObjectName("fasta_rb")
        self.verticalLayout2.addWidget(self.fasta_rb)
        self.UniProtSearchFasta_pb = QtWidgets.QPushButton(self.verticalLayoutWidget2)
        self.UniProtSearchFasta_pb.setObjectName("UniProtSearchFasta_pb")
        self.verticalLayout2.addWidget(self.UniProtSearchFasta_pb) 
        self.outputScrollArea = QtWidgets.QScrollArea(self.verticalLayoutWidget2)
        self.outputScrollArea.setObjectName("outputScrollArea")
        self.outputScrollArea.setWidgetResizable(False)
        self.verticalLayout2.addWidget(self.outputScrollArea)
        self.outputScrollAreaWidget = QtWidgets.QWidget(self.verticalLayoutWidget2)
        self.outputScrollAreaWidget.setObjectName("outputScrollAreaWidget")
        self.Uniprot_output_lb = QtWidgets.QLabel(self.outputScrollAreaWidget)
        self.Uniprot_output_lb.setObjectName("Uniprot_output_lb")
        self.outputScrollArea.setWidget(self.outputScrollAreaWidget)
        self.tabWidget.addTab(self.tab_FindProt, "")
        
        self.tab_inSilicoDigestion = QtWidgets.QWidget()
        self.tab_inSilicoDigestion.setObjectName("tab_inSilicoDigestion")
        self.verticalLayoutWidget3 = QtWidgets.QWidget(self.tab_inSilicoDigestion)
        self.verticalLayoutWidget3.setGeometry(self.LayoutWidgetGeometry)
        self.verticalLayoutWidget3.setObjectName("verticalLayoutWidget3")
        self.verticalLayout3 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget3)
        self.verticalLayout3.setContentsMargins(0, 0, 0, 0)
        self.label3 = QtWidgets.QLabel()
        self.label3.setAlignment(QtCore.Qt.AlignCenter)
        self.label3.setObjectName("label3")
        self.verticalLayout3.addWidget(self.label3)
        self.browse_sequence_pb = QtWidgets.QPushButton()
        self.browse_sequence_pb.setObjectName("browse_sequence_pb")
        self.verticalLayout3.addWidget(self.browse_sequence_pb)
        self.sequence_textBox = QtWidgets.QTextEdit()
        self.sequence_textBox.setObjectName("sequence_textBox")
        self.verticalLayout3.addWidget(self.sequence_textBox)
        self.enzyme_layout = QtWidgets.QHBoxLayout()
        self.enzyme_layout.setObjectName("enzyme_layout")
        self.enzyme_label = QtWidgets.QLabel()
        self.enzyme_label.setObjectName("enzyme_label")
        self.enzyme_layout.addWidget(self.enzyme_label)
        self.enzyme_cb = QtWidgets.QComboBox()
        self.enzyme_cb.setObjectName("enzyme_cb")
        self.enzyme_layout.addWidget(self.enzyme_cb)
        self.digestion_pb = QtWidgets.QPushButton()
        self.digestion_pb.setObjectName("digestion_pb")
        self.enzyme_layout.addWidget(self.digestion_pb)
        self.verticalLayout3.addLayout(self.enzyme_layout)
        self.tabWidget.addTab(self.tab_inSilicoDigestion, "")
        
        self.tab_mzPeakShape = QtWidgets.QWidget()
        self.tab_mzPeakShape.setObjectName("tab_mzPeakShape")      
        self.verticalLayoutWidget4 = QtWidgets.QWidget(self.tab_mzPeakShape)
        self.verticalLayoutWidget4.setGeometry(self.LayoutWidgetGeometry)
        self.verticalLayoutWidget4.setObjectName("verticalLayoutWidget4")
        self.verticalLayout4 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget4)
        self.verticalLayout4.setContentsMargins(0,0,0,0)       
        self.label4 = QtWidgets.QLabel()
        self.label4.setAlignment(QtCore.Qt.AlignCenter)
        self.label4.setObjectName("label4")
        self.verticalLayout4.addWidget(self.label4)
        
        self.mzNameLayout = QtWidgets.QHBoxLayout()
        self.mzNameLayout.setObjectName("mzNameLayout")

        self.name_lb = QtWidgets.QLabel()
        self.name_lb.setObjectName("name_lb")
        self.mzNameLayout.addWidget(self.name_lb)       
        self.name_lineEdit = QtWidgets.QLineEdit()
        self.name_lineEdit.setObjectName("name_lineEdit")
        self.mzNameLayout.addWidget(self.name_lineEdit)
        self.verticalLayout4.addLayout(self.mzNameLayout)
        
        self.mzOptionsLayout = QtWidgets.QHBoxLayout()
        self.mzOptionsLayout.setObjectName("mzOptionsLayout")
        
        self.charge_lb = QtWidgets.QLabel()
        self.charge_lb.setObjectName("charge_lb")
        self.mzOptionsLayout.addWidget(self.charge_lb)
        self.charge_sb = QtWidgets.QSpinBox()
        self.charge_sb.setObjectName("charge_sb")
        self.mzOptionsLayout.addWidget(self.charge_sb)
        self.adduct_lb = QtWidgets.QLabel()
        self.adduct_lb.setObjectName("adduct_lb")
        self.mzOptionsLayout.addWidget(self.adduct_lb)
        self.adduct_cb = QtWidgets.QComboBox()
        self.adduct_cb.setObjectName("adduct_cb")
        self.mzOptionsLayout.addWidget(self.adduct_cb)
        self.mod_lb = QtWidgets.QLabel()
        self.mod_lb.setObjectName("mod_lb")
        self.mzOptionsLayout.addWidget(self.mod_lb)
        self.mod_cb = QtWidgets.QComboBox()
        self.mod_cb.setObjectName("mod_cb")
        self.mzOptionsLayout.addWidget(self.mod_cb)
        self.verticalLayout4.addLayout(self.mzOptionsLayout)  
        
        self.textEditSequence_mz = QtWidgets.QTextEdit()
        self.textEditSequence_mz.setObjectName("textEditSequence_mz")
        self.verticalLayout4.addWidget(self.textEditSequence_mz)
        self.mzShape_pb = QtWidgets.QPushButton()
        self.mzShape_pb.setObjectName("mzShape_pb")
        self.verticalLayout4.addWidget(self.mzShape_pb)
        #PUT STUFF HERE
        
        self.tabWidget.addTab(self.tab_mzPeakShape, "")
        
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 800, 30))
        self.menubar.setObjectName("menubar")
        self.menuFile = QtWidgets.QMenu(self.menubar)
        self.menuFile.setObjectName("menuFile")
        self.menuView = QtWidgets.QMenu(self.menubar)
        self.menuView.setObjectName("menuEdit")
        self.menuHelp = QtWidgets.QMenu(self.menubar)
        self.menuHelp.setObjectName("menuHelp")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.actionOpen = QtWidgets.QAction(MainWindow)
        self.actionOpen.setObjectName("actionOpen")
        self.actionOpen.setShortcut("Ctrl+O")
        self.actionSave = QtWidgets.QAction(MainWindow)
        self.actionSave.setObjectName("actionSave")
        self.actionSave.setShortcut("Ctrl+S")
        self.actionSave_as = QtWidgets.QAction(MainWindow)
        self.actionSave_as.setObjectName("actionSave_as")
        self.actionSave_as.setShortcut("Ctrl+Shift+S")
        self.actionLight = QtWidgets.QAction(MainWindow)
        self.actionLight.setObjectName("actionLight")
        self.actionLight.setShortcut("Ctrl+Shift+L")
        self.actionDark = QtWidgets.QAction(MainWindow)
        self.actionDark.setObjectName("actionDark")
        self.actionDark.setShortcut("Ctrl+Shift+D")
        self.actionManual = QtWidgets.QAction(MainWindow)
        self.actionManual.setObjectName("actionManual")
        self.actionManual.setShortcut("Ctrl+H")
        self.menuFile.addAction(self.actionOpen)
        self.menuFile.addAction(self.actionSave)
        self.menuFile.addAction(self.actionSave_as)
        self.menuView.addAction(self.actionLight)
        self.menuView.addAction(self.actionDark)
        self.menuHelp.addAction(self.actionManual)
        self.menubar.addAction(self.menuFile.menuAction())
        self.menubar.addAction(self.menuView.menuAction())
        self.menubar.addAction(self.menuHelp.menuAction())
        self.centralwidget.adjustSize()
        self.retranslateUi(MainWindow)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "SSPC Lab Toolkit"))
        self.label.setText(_translate("MainWindow", "How many components do you have in your sample?"))
        self.label.setFont(self.mainFont)
        self.spinBox.setFont(self.mainFont)
        self.pushButton.setText(_translate("MainWindow", "Find Weight"))
        self.pushButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.pushButton.setFont(self.mainFont)
        self.pushButton.clicked.connect(lambda: self.openFindWeightForm(int(self.spinBox.value())))
        self.pushButton.setShortcut('Return')
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_FindWeight), _translate("MainWindow", "Find Weight"))
        
        self.label2.setText("Enter a UniProt protein ID:")
        self.label2.setFont(self.mainFont)
        self.prot_lineEdit.setFont(self.mainFont)
        self.prot_lineEdit.setPlaceholderText("UniProt ID")
        self.text_rb.setText("search protein information")
        self.text_rb.setFont(self.mainFont)
        self.fasta_rb.setText("search FASTA sequence")
        self.fasta_rb.setFont(self.mainFont)
        self.fasta_rb.setChecked(True)
        self.UniProtSearchFasta_pb.setText(_translate("MainWindow", "Find Protein"))
        self.UniProtSearchFasta_pb.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.UniProtSearchFasta_pb.setFont(self.mainFont)
        self.UniProtSearchFasta_pb.clicked.connect(lambda: self.UniProt_Search(isFasta = self.fasta_rb.isChecked()))
        self.UniProtSearchFasta_pb.setShortcut('Return')
        self.Uniprot_output_lb.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse)
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_FindProt), _translate("MainWindow", "Find Protein"))
        
        self.label3.setText("Enter a protein sequence or browse a FASTA file:")
        self.label3.setFont(self.mainFont)
        self.sequence_textBox.setFont(self.fastaFont)
        self.sequence_textBox.setPlaceholderText("Enter a Protein sequence here...\n"+\
                                                 "If several sequences, enter them in FASTA format:\n\n"+\
                                                 '>"Accession1"\n"SEQUENCE1"\n\n>"Accession2"\n"SEQUENCE2"\n\netc...')
        self.browse_sequence_pb.setText("Browse FASTA")
        self.browse_sequence_pb.setFont(self.mainFont)
        self.browse_sequence_pb.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.browse_sequence_pb.clicked.connect(lambda:self.browseFASTASequence())
        self.enzymeList = ["None","Trypsin","FabULOUS SpeB", "FabRICATOR IdeS",
                           "GingisKHAN Kgp"]
        self.enzymeList.sort()
        self.enzyme_label.setText("Choose an enzyme:")
        self.enzyme_label.setFont(self.mainFont)
        self.enzyme_cb.addItems(self.enzymeList)
        self.enzyme_cb.setFont(self.mainFont)
        self.enzyme_cb.setCurrentText("None")
        self.digestion_pb.setText("Launch digestion")
        self.digestion_pb.setFont(self.mainFont)
        self.digestion_pb.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.digestion_pb.clicked.connect(
            lambda:self.ProteinDigestion(self.enzyme_cb.currentText()))
        self.tabWidget.setTabText(
            self.tabWidget.indexOf(self.tab_inSilicoDigestion),
            _translate("MainWindow", "in Silico Digestion"))
        self.label4.setText("Enter a peptide sequence and set the parameters"\
                            +" to display a theoretical peak massif")
        self.label4.setFont(self.mainFont)
        self.name_lb.setText("graph name:")
        self.name_lb.setFont(self.mainFont)
        self.name_lb.setAlignment(QtCore.Qt.AlignRight)
        self.name_lineEdit.setPlaceholderText('User Input')
        self.name_lineEdit.setFont(self.mainFont)
        self.charge_lb.setText("charge:")
        self.charge_lb.setAlignment(QtCore.Qt.AlignRight)
        self.charge_lb.setFont(self.mainFont)
        self.charge_sb.setMinimum(1)
        self.charge_sb.setMaximum(10)
        self.charge_sb.setFont(self.mainFont)
        self.adduct_lb.setText("adduct:")
        self.adduct_lb.setAlignment(QtCore.Qt.AlignRight)
        self.adduct_lb.setFont(self.mainFont)
        self.adductItems = ('Proton','Lithium','Ammonium','Sodium','Potassium')
        self.adduct_cb.addItems(self.adductItems)
        self.adduct_cb.setFont(self.mainFont)
        self.mod_lb.setText("modification")
        self.mod_lb.setAlignment(QtCore.Qt.AlignRight)
        self.mod_lb.setFont(self.mainFont)
        self.modItems = ("None", "Carbamidomethylation")
        self.mod_cb.addItems(self.modItems)
        self.mod_cb.setFont(self.mainFont) 
        self.textEditSequence_mz.setPlaceholderText("Put a protein sequence")
        self.textEditSequence_mz.setFont(self.fastaFont)
        
        self.mzShape_pb.setText("Display m/z massif shape")
        self.mzShape_pb.setFont(self.mainFont)
        self.mzShape_pb.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        
        self.mzShape_pb.clicked.connect(lambda:self.openPeakShapeWindow(
            name = self.name_lineEdit.text(),
            sequence= self.textEditSequence_mz.toPlainText(),
            charge= self.charge_sb.value(),
            adduct= self.adduct_cb.currentText(),
            mod= self.mod_cb.currentText()))
        
        #PUT STUFF FROM TAB4 HERE
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_mzPeakShape), _translate("MainWindow", "m/z Peak Shape"))
        
        self.tabWidget.setFont(self.mainFont)
        self.menuFile.setTitle(_translate("MainWindow", "File"))
        self.menuView.setTitle(_translate("MainWindow", "View"))
        self.menuHelp.setTitle(_translate("MainWindow", "Help"))
        self.actionOpen.setText(_translate("MainWindow", "Open"))
        self.actionOpen.triggered.connect(lambda:self.openFile(self.tabWidget.currentWidget().objectName()))
        self.actionSave.setText(_translate("MainWindow", "Save"))
        self.actionSave.triggered.connect(lambda: self.save(self.tabWidget.currentWidget().objectName()))
        self.actionSave_as.setText(_translate("MainWindow", "Save as ..."))
        self.actionSave_as.triggered.connect(lambda: self.saveAs(self.tabWidget.currentWidget().objectName()))
        self.actionLight.setText(_translate("MainWindow", "Light Mode"))
        self.actionLight.triggered.connect(lambda:self.switchColor('light'))
        self.actionDark.setText(_translate("MainWindow", "Dark Mode"))
        self.actionDark.triggered.connect(lambda:self.switchColor('dark'))
        self.actionManual.setText(_translate("MainWindow", "Manual"))
        self.actionManual.triggered.connect(lambda: self.openManual()) 
        
        MainWindow.setStyleSheet(qdarktheme.load_stylesheet(
            f"{'light' if self.lightMode == True else 'dark'}"))

    def openFindWeightForm(self, compnbs):
        self.statusbar.showMessage("open second Window", 2000)
        self.window = QtWidgets.QWidget()
        self.window.setStyleSheet(qdarktheme.load_stylesheet(f"{'light' if self.lightMode == True else 'dark'}"))
        self.ui = secondWindow_v2.Ui_Form()
        self.ui.setupUi(self.window)
        self.window.setWindowIcon(self.mainIcon) 
        self.ui.label_name.setFont((self.mainFont))
        self.ui.label_pb_convert.setFont((self.mainFont))
        self.ui.label_MW.setFont((self.mainFont))
        self.ui.label_C.setFont((self.mainFont))
        self.ui.label_C_unit.setFont((self.mainFont))
        self.ui.label_V.setFont((self.mainFont))
        self.ui.label_V_unit.setFont((self.mainFont))
        self.ui.label_pb_calc.setFont((self.mainFont))
        self.ui.label_mass.setFont((self.mainFont))
        self.ui.OpenBtn.clicked.connect(lambda: self.openFile("tab_FindWeight"))
        self.ui.saveBtn.clicked.connect(lambda: self.save("tab_FindWeight"))
        self.ui.saveAsBtn.clicked.connect(lambda: self.saveAs("tab_FindWeight"))
        self.ui.LightBtn.clicked.connect(lambda: self.switchColor('light'))
        self.ui.DarkBtn.clicked.connect(lambda: self.switchColor('dark'))
        self.ui.HelpBtn.clicked.connect(lambda: self.openManual())
        self.comp_lineEdits = {}
        self.comp_MW_pbs = {}
        self.label_MWs = {}
        self.C_lineEdits = {}
        self.C_combos = {}
        self.V_lineEdits = {}
        self.V_combos = {}
        self.calc_pbs = {}
        self.mass_lbs = {}
        
        nb_list = []
        for compnb in range(compnbs):
            nb_list.append(compnb+1)
            self.comp_lineEdit = QtWidgets.QLineEdit(self.ui.gridLayoutWidget)
            self.comp_lineEdit.setObjectName(f"comp_lineEdit_{compnb+1}")
            self.comp_lineEdit.setFont(self.mainFont)
            self.ui.grid_comp.addWidget(self.comp_lineEdit, compnb+1, 0)
            self.comp_lineEdits[f"comp_lineEdit_{compnb+1}"] = self.comp_lineEdit
            
            self.comp_MW_pb = QtWidgets.QPushButton(self.ui.gridLayoutWidget)
            self.comp_MW_pb.setObjectName(f"comp_MW_pb_{compnb+1}")
            self.comp_MW_pb.setText("search")
            self.comp_MW_pb.clicked.connect(lambda checked, nb = compnb+1: self.PUGREST_Search(nb))
            self.comp_MW_pb.setFont(self.mainFont)
            self.comp_MW_pb.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            self.ui.grid_comp.addWidget(self.comp_MW_pb, compnb+1, 1)
            self.comp_MW_pbs[f"comp_MW_pb_{compnb+1}"] = self.comp_MW_pb
            
            self.label_MW = QtWidgets.QLabel(self.ui.gridLayoutWidget)
            self.label_MW.setObjectName(f"label_MW_{compnb+1}")
            self.label_MW.setAlignment(QtCore.Qt.AlignCenter)
            self.label_MW.setFont(self.mainFont)
            self.ui.grid_comp.addWidget(self.label_MW, compnb+1, 2)
            self.label_MWs[f"label_MW_{compnb+1}"] = self.label_MW
            
            self.C_lineEdit = QtWidgets.QLineEdit(self.ui.gridLayoutWidget)
            self.C_lineEdit.setObjectName(f"C_lineEdit_{compnb+1}")
            self.C_lineEdit.setFont(self.mainFont)
            self.ui.grid_comp.addWidget(self.C_lineEdit, compnb+1, 3)
            self.C_lineEdit.setValidator(QtGui.QDoubleValidator())
            self.C_lineEdits[f"C_lineEdit_{compnb+1}"] = self.C_lineEdit
            
            self.C_combo = QtWidgets.QComboBox(self.ui.gridLayoutWidget)
            self.C_combo.setObjectName(f"C_combo_{compnb+1}")
            self.C_combo.setFont(self.mainFont)
            self.ui.grid_comp.addWidget(self.C_combo, compnb+1, 4)
            self.C_combo.insertItems(0, ('mM','M'))
            self.C_combos[f"C_combo_{compnb+1}"] = self.C_combo
            
            self.V_lineEdit = QtWidgets.QLineEdit(self.ui.gridLayoutWidget)
            self.V_lineEdit.setObjectName(f"V_lineEdit_{compnb+1}")
            self.V_lineEdit.setFont(self.mainFont)
            self.ui.grid_comp.addWidget(self.V_lineEdit, compnb+1, 5)
            self.V_lineEdit.setValidator(QtGui.QDoubleValidator())
            self.V_lineEdits[f"V_lineEdit_{compnb+1}"] = self.V_lineEdit
            
            self.V_combo = QtWidgets.QComboBox(self.ui.gridLayoutWidget)
            self.V_combo.setObjectName(f"V_combo_{compnb+1}")
            self.V_combo.setFont(self.mainFont)
            self.ui.grid_comp.addWidget(self.V_combo, compnb+1, 6)
            self.V_combo.insertItems(0, ('L','mL', '\u00b5L'))
            self.V_combos[f"V_combo_{compnb+1}"] = self.V_combo
            
            self.calc_pb = QtWidgets.QPushButton(self.ui.gridLayoutWidget)
            self.calc_pb.setObjectName(f"calc_pb_{compnb+1}")
            self.calc_pb.setText("Calculate")
            self.calc_pb.setFont(self.mainFont)
            self.calc_pb.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
            self.calc_pb.clicked.connect(lambda checked, nb = compnb+1: self.calcMass(nb))
            self.ui.grid_comp.addWidget(self.calc_pb, compnb+1, 7)
            self.calc_pbs[f"calc_pb_{compnb+1}"] = self.comp_MW_pb
            
            self.mass_lb = QtWidgets.QLabel(self.ui.gridLayoutWidget)
            self.mass_lb.setObjectName(f"mass_lb_{compnb+1}")
            self.mass_lb.setAlignment(QtCore.Qt.AlignCenter)
            self.mass_lb.setFont(self.mainFont)
            self.ui.grid_comp.addWidget(self.mass_lb, compnb+1, 8)
            self.mass_lbs[f"mass_lb_{compnb+1}"] = self.mass_lb
        
        self.ui.gridLayoutWidget.adjustSize()
        width =  self.ui.gridLayoutWidget.frameGeometry().width()
        height = self.ui.gridLayoutWidget.frameGeometry().height()
        self.ui.scrollArea.setGeometry(0, 0, width, height)
        self.window.show()
    
    def openTextWindow(self, text):
        self.statusbar.showMessage("open second Window", 2000)
        self.textWindow = QtWidgets.QWidget()
        self.textWindow.setWindowIcon(self.mainIcon)
        self.UP_ui = UniProtWindow.Ui_UniProtForm()
        self.UP_ui.setupUi(self.textWindow)
        activeTabWhenOpened = self.tabWidget.currentWidget().objectName()
        self.UP_ui.OpenBtn.clicked.connect(lambda: self.openFile(activeTabWhenOpened))
        self.UP_ui.saveBtn.clicked.connect(lambda: self.save(activeTabWhenOpened))
        self.UP_ui.saveAsBtn.clicked.connect(lambda: self.saveAs(activeTabWhenOpened))
        self.UP_ui.LightBtn.clicked.connect(lambda: self.switchColor('light'))
        self.UP_ui.DarkBtn.clicked.connect(lambda: self.switchColor('dark'))
        self.UP_ui.HelpBtn.clicked.connect(lambda: self.openManual())
        self.UP_ui.UniProtTextLabel.setText(text)
        self.textWindow.setStyleSheet(qdarktheme.load_stylesheet(f"{'light' if self.lightMode == True else 'dark'}"))
        self.UP_ui.UniProtTextLabel.setFont(self.fastaFont)
        self.UP_ui.UniProtTextLabel.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse)
        self.UP_ui.UniProtTextLabel.setWordWrap(True)
        self.UP_ui.UniProtTextLabel.setCursor(QtGui.QCursor(QtCore.Qt.IBeamCursor))
        self.UP_ui.UniProtTextLabel.adjustSize()
        self.UP_ui.UniProtScrollAreaWidget.adjustSize()
        self.textWindow.show()
    
    def unknownCharactersError(self, sequence):
            Charwarningbox = QtWidgets.QMessageBox()
            Charwarningbox.setIcon(QtWidgets.QMessageBox.Warning)
            Charwarningbox.setWindowIcon(self.mainIcon)
            Charwarningbox.setWindowTitle('Something wrong with the sequence')
            Charwarningbox.setStyleSheet(qdarktheme.load_stylesheet(f"{'light' if self.lightMode == True else 'dark'}"))
            Charwarningbox.setText("One or several characters in the sequence do not correspond to amino acids.\n"\
                                   +f"{[x for x in sequence if x not in aminoDict['Code']]}")    
            Charwarningbox.exec_()
            self.statusbar.showMessage('peak massif display aborted', 2000)
    
    def openPeakShapeWindow(self, name, sequence, charge, adduct, mod):
        
        sequence = sequence.strip()
        sequence = sequence.replace("\n", "")
        sequence = sequence.upper()
        readable = True
        if name == "":
            name = "User Input"
        for char in sequence:
            if char not in aminoDict['Code']:
                readable = False
        
        if readable == True:
            self.statusbar.showMessage("open Peak Shape Window")
            self.peakShapeWindow = QtWidgets.QWidget() 
            self.PeakShape_ui = PlotWindow.Ui_PlotForm()
            activeTabWhenOpened = self.tabWidget.currentWidget().objectName()

            self.peakDataset = mzPeakMassif(
                accession = name,
                sequence = sequence,
                charge = charge,
                adduct = adduct,
                modification = mod)
            
            self.PeakShape_ui.setupUi(self.peakShapeWindow,
                accession=name,
                data= self.peakDataset,
                charge = charge,
                adduct = adduct,
                modification = mod)
            
            self.PeakShape_ui.OpenBtn.clicked.connect(lambda: self.openFile(activeTabWhenOpened))
            self.PeakShape_ui.saveBtn.clicked.connect(lambda: self.save(activeTabWhenOpened))
            self.PeakShape_ui.saveAsBtn.clicked.connect(lambda: self.saveAs(activeTabWhenOpened))
            self.PeakShape_ui.LightBtn.clicked.connect(lambda: self.switchColor('light'))
            self.PeakShape_ui.DarkBtn.clicked.connect(lambda: self.switchColor('dark'))
            self.PeakShape_ui.HelpBtn.clicked.connect(lambda: self.openManual())
        
            self.peakShapeWindow.setWindowIcon(self.mainIcon)
            self.peakShapeWindow.setStyleSheet(qdarktheme.load_stylesheet(f"{'light' if self.lightMode == True else 'dark'}"))
            self.peakShapeWindow.setWindowTitle("Peak massif")
            self.peakShapeWindow.show()
        
        else: self.unknownCharactersError(sequence)

    
    def MemoryWarning(self, output_location, file):
        if  len(file) < 100000:
            MemoryWarning = QtWidgets.QMessageBox()
            MemoryWarning.setStyleSheet(qdarktheme.load_stylesheet(f"{'light' if self.lightMode == True else 'dark'}"))
            MemoryWarning.setWindowIcon(self.mainIcon)
            MemoryWarning.setWindowTitle("Memory Warning")
            MemoryWarning.setIcon(QtWidgets.QMessageBox.Warning)
            MemoryWarning.setText("Your input contain more than 10,000 characters:\n"+\
                                  f"{len(file)} characters.\n"+\
                                  "Opening this file might slow down the app.\n"+\
                                  "Continue the upload anyway?")
            MemoryWarning.setStandardButtons(QtWidgets.QMessageBox.Yes|QtWidgets.QMessageBox.No)
            def ButtonClicked(button):
                if button.text() == '&Yes': output_location.setText(file) 
                else: pass 
            MemoryWarning.buttonClicked.connect(ButtonClicked)
            MemoryWarning.exec_()
        else:
            MemoryCritical = QtWidgets.QMessageBox()
            MemoryCritical.setStyleSheet(qdarktheme.load_stylesheet(f"{'light' if self.lightMode == True else 'dark'}"))
            MemoryCritical.setWindowIcon(self.mainIcon)
            MemoryCritical.setWindowTitle("Memory Critical")
            MemoryCritical.setIcon(QtWidgets.QMessageBox.Critical)
            MemoryCritical.setText("Your input contain more than 100,000 characters:\n"+\
                                  f"{len(file)} characters.\n"+\
                                  "Opening this file will surely freeze the app.\n"+\
                                  "This app is not meant to open large datafiles such as complete proteome.\n"+\
                                  "Continue the upload anyway?")
            MemoryCritical.setStandardButtons(QtWidgets.QMessageBox.Yes|QtWidgets.QMessageBox.No)
            def ButtonClicked(button):
                if button.text() == '&Yes': output_location.setText(file) 
                else: pass 
            MemoryCritical.buttonClicked.connect(ButtonClicked)
            MemoryCritical.exec_()     
    
    def browseFASTASequence(self):
        self.browseFASTA = QtWidgets.QFileDialog().getOpenFileName(
            MainWindow, 
            'Browse FASTA sequence',
            '',
            'FASTA file (*.fasta)')
        if self.browseFASTA[0] == '':
            delattr(self, 'browseFASTA')
            
        else:   
            if self.browseFASTA[1] == "FASTA file (*.fasta)":
                with open(self.browseFASTA[0], 'r') as openFile:
                    fastaFile = openFile.read()
                
                if len(fastaFile) < 10000:  
                    self.sequence_textBox.setText(fastaFile)
                    
                else: self.MemoryWarning(self.sequence_textBox, fastaFile)             
                                    
            else:self.fileNotRecognized()  
    
    def Error(self, http, error, ID, db_name):
        Prot_warningbox1 = QtWidgets.QMessageBox()
        Prot_warningbox1.setStyleSheet(qdarktheme.load_stylesheet(f"{'light' if self.lightMode == True else 'dark'}"))
        Prot_warningbox1.setWindowIcon(self.mainIcon)
        Prot_warningbox1.setWindowTitle(f"{error}")     
        if http == True:
            Prot_warningbox1.setIcon(QtWidgets.QMessageBox.Warning)
            Prot_warningbox1.setText(f"{ID} cannot be found in {db_name}.\n\n"\
                            "Please, verify if there is no typo.")
        else:
            Prot_warningbox1.setIcon(QtWidgets.QMessageBox.Critical)
            Prot_warningbox1.setText(f"Your computer cannot access to {db_name}.\n\n"\
                            "Please, verify your internet connexion.")
  
        Prot_warningbox1.exec_()
        self.Uniprot_output_lb.setText(str(error))
        
        if http == True:
            self.statusbar.showMessage(f"{ID} not found: {error}", 4000)
        else:
            self.statusbar.showMessage(f"cannot access to {db_name}: {error}", 4000)
    
    def UniProt_Search(self, isFasta):
        UniProtID = str(self.prot_lineEdit.text())
        UniProtID = UniProtID.strip()
        UniProtID = UniProtID.replace(' ', '')
        self.statusbar.showMessage(f"Looking for {UniProtID} in UniProt...")
        if isFasta == True:
            try:
                UniProt_Request1 = f'https://www.uniprot.org/uniprot/{UniProtID}.fasta'
                with url(UniProt_Request1) as request:
                    response1 = request.read().decode('utf-8')
                                            
                self.Uniprot_output_lb.setText(response1)
                self.Uniprot_output_lb.setFont(self.fastaFont)
                self.Uniprot_output_lb.setWordWrap(True)
                self.Uniprot_output_lb.setCursor(QtGui.QCursor(QtCore.Qt.IBeamCursor))
                self.Uniprot_output_lb.adjustSize()
                self.outputScrollAreaWidget.adjustSize()
                self.statusbar.showMessage(f"{UniProtID} found", 2000)
                
            except HTTPError as e: self.Error(http = True, error = e, ID = UniProtID, db_name = "UniProt")
            except URLError as e: self.Error(http = False, error = e, ID = UniProtID, db_name = "UniProt")

        else:
            try:            
                UniProt_Request2 = f'https://www.uniprot.org/uniprot/{UniProtID}.txt'
                with url(UniProt_Request2) as request:
                    response2 = request.read().decode('utf-8')
                self.openTextWindow(response2)
                self.statusbar.showMessage(f"{UniProtID} found", 2000)

            except HTTPError as e: self.Error(http = True, error = e, ID = UniProtID, db_name = "UniProt")
            except URLError as e: self.Error(http = False, error = e, ID = UniProtID, db_name = "UniProt")

    def PUGREST_Search(self, nb):
        name = str(self.comp_lineEdits[f"comp_lineEdit_{nb}"].text())
        name = name.strip()
        name = name.replace(' ', '-')
        self.statusbar.showMessage(f"looking for {name} in PubChem...")
        try:
            PUG_Request = f'https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{name}/property/MolecularWeight/TXT' 
            with url(PUG_Request) as request:
                self.response = request.read().decode('utf-8')

                try:
                    self.output = float(self.response)
                    #if several MW exist for a component (i.e. amino-acids in their acidic, zwitterion and alkaline form)
                    # response will be a string of this form: 'a\nb\nc\n'... which cannot be converted as a float
                    # the exception will convert response in a list and the user will have to choose one value
                except: 
                    self.response = self.response.split('\n')
                    self.response.remove('')
        
                   #if several items in the list have the same value, creating a set will remove the iterations
                    self.response = set(self.response)
                    self.response = list(self.response)
                    
                    if len(self.response) >2:
                        ms_box1 = QtWidgets.QMessageBox()
                        ms_box1.setIcon(QtWidgets.QMessageBox.Information)
                        ms_box1.setWindowTitle('Several MW found with PUGREST')
                        ms_box1.setWindowIcon(self.mainIcon)
                        ms_box1.setStyleSheet(qdarktheme.load_stylesheet(f"{'light' if self.lightMode == True else 'dark'}"))
                        ms_box1.setText('Several MW values were proposed for your component.\n')
                        ms_box1.exec_()
                        self.MW_window = QtWidgets.QWidget()
                        self.MW_ui = choose_MW.Ui_MW_list_Form()
                        self.MW_ui.setupUi(self.MW_window)
                        self.MW_window.setWindowIcon(self.mainIcon)
                        self.MW_window.setStyleSheet(qdarktheme.load_stylesheet(f"{'light' if self.lightMode == True else 'dark'}"))
                        self.MW_ui.label_choose.setFont(self.mainFont)
                        QText_str = ''
                        for index, mw in enumerate(self.response):
                            QText_str += f'{index}:\t{mw}\n'
                        self.MW_ui.MW_list_lb.setText(QText_str)
                        self.MW_ui.MW_list_lb.setFont(self.mainFont)
                        self.MW_ui.spinBox_indexes.setMaximum(len(self.response)-1)
                        self.MW_ui.spinBox_indexes.setFont(self.mainFont)
                        self.chosen_index = 0
                        self.MW_ui.OK_pb.setFont(self.mainFont)
                        self.MW_ui.OK_pb.clicked.connect(lambda: self.chosenIndex(nb))
                        self.MW_window.show()
                        self.output = self.response[self.chosen_index]
                        
                    else:
                        self.output = self.response[0]
                self.statusbar.showMessage(f"{name} found! status code: {request.status}", 2000)
        except HTTPError as e: self.Error(http = True, error = e, ID = name, db_name = "PubChem")   
        except URLError as e: self.Error(http = False, error = e, ID = name, db_name = "PubChem")
                    
        self.label_MWs[f"label_MW_{nb}"].setText(str(self.output))
    
    def ProteinDigestion(self, enzyme):
        self.statusbar.showMessage(enzyme, 2000)
        
        if enzyme == "Trypsin":
            fragments = InSilicoDigestion.Trypsin(self.sequence_textBox.toPlainText())   
        elif enzyme == "FabULOUS SpeB":
            fragments = InSilicoDigestion.FabULOUS_SpeB(self.sequence_textBox.toPlainText())
        elif enzyme == "FabRICATOR IdeS":
            fragments = InSilicoDigestion.FabRICATOR_IdeS(self.sequence_textBox.toPlainText())
        elif enzyme == "GingisKHAN Kgp":
            fragments = InSilicoDigestion.GingisKHAN_Kgp(self.sequence_textBox.toPlainText())
        elif enzyme == "None":
            fragments = InSilicoDigestion.NoDigestion(self.sequence_textBox.toPlainText())
        
        self.digestionPlainText = ''
        line = 0
        displayWarningBox = False
        for ident, sequence in fragments:
            line += 1
            try: MW = peptideMW(sequence)
            except: 
                displayWarningBox = True
                MW = 'unknown'
                
            sequence_50char = '\n'.join(sequence[i:i+50] for i in range(0, len(sequence), 50))
            self.digestionPlainText = self.digestionPlainText + f">{line}_{ident}_MW:{MW}_Da\n{sequence_50char}\n\n"

        if displayWarningBox == True:
                MWwarningbox = QtWidgets.QMessageBox()
                MWwarningbox.setWindowIcon(self.mainIcon)
                MWwarningbox.setIcon(QtWidgets.QMessageBox.Warning)
                MWwarningbox.setWindowTitle('Something wrong with MW calculation')
                MWwarningbox.setStyleSheet(qdarktheme.load_stylesheet(f"{'light' if self.lightMode == True else 'dark'}"))
                MWwarningbox.setText("Some amino acid in some of the sequences were unreadable.\n"+\
                                     "Cannot calculate their MW.\n"+\
                                     'this can happen it you have "X" in your sequence,\n'+\
                                     'corresponding to an unknown amino acid')
                MWwarningbox.exec_()
        
        self.openTextWindow(self.digestionPlainText)       

    def calcMass(self, nb):
        try:
            self.statusbar.showMessage('Calculate the mass...')
            M = float(self.label_MWs[f'label_MW_{nb}'].text())
            C = float(self.C_lineEdits[f'C_lineEdit_{nb}'].text())
            finalV = float(self.V_lineEdits[f"V_lineEdit_{nb}"].text())
            C_unit = 0.001
            V_unit = 1
            
            if str(self.C_combos[f"C_combo_{nb}"].currentText()) == "mM":   
                C_unit = 0.001
            elif str(self.C_combos[f"C_combo_{nb}"].currentText()) == "M":
                C_unit = 1
            else:
                C_unit = 0.001
    
                
            if str(self.V_combos[f"V_combo_{nb}"].currentText()) =="L":
                V_unit = 1
            elif str(self.V_combos[f"V_combo_{nb}"].currentText()) =="mL":
                V_unit = 0.001 
            elif str(self.V_combos[f"V_combo_{nb}"].currentText()) =="\u00b5L":
                V_unit = 0.000001
            else:
                V_unit = 1
                
            mass = C*M*finalV*C_unit*V_unit
            mass_unit = "g"
            if mass >= 1000:
                mass = mass * 0.001
                mass_unit = "kg"
            elif 1 <= mass < 1000:
                mass = mass
                mass_unit = "g"
            elif 0.001 <= mass < 1:
                mass = mass * 1000
                mass_unit = "mg"
            elif mass < 0.001:
                mass = mass * 1000000
                mass_unit = "\u00b5g"
                
            self.statusbar.showMessage('Mass calculated', 2000)
            self.mass_lbs[f"mass_lb_{nb}"].setText(f"{round(mass, 3)} {mass_unit}")
            
        except:
            warningbox2 = QtWidgets.QMessageBox()
            warningbox2.setIcon(QtWidgets.QMessageBox.Warning)
            warningbox2.setWindowIcon(self.mainIcon)
            warningbox2.setWindowTitle('Something wrong with calculation')
            warningbox2.setStyleSheet(qdarktheme.load_stylesheet(f"{'light' if self.lightMode == True else 'dark'}"))
            warningbox2.setText("One or several parameters have not been setup correctly.\n")
            warningbox2.exec_()
            self.statusbar.showMessage('Calculation aborted', 2000)
            self.mass_lbs[f"mass_lb_{nb}"].setText("Error")
                            
    def chosenIndex(self, nb):
        self.chosen_index = int(self.MW_ui.spinBox_indexes.text())
        self.output = self.response[self.chosen_index]
        self.label_MWs[f"label_MW_{nb}"].setText(str(self.output))
        self.MW_window.close()
    
    def writeTXT(self, filename, tab):
        
        if tab == "tab_FindWeight":
            
            self.statusbar.showMessage('Saving Find Weight data as .txt')
            
            with open(filename[0], 'w') as saveFile:
                saveFile.write("Name\tMW\tConcentration\tConcentration unit\tVolume\tVolume unit\tmass\n")
                for i in range(len(self.comp_lineEdits)):
                    saveFile.write( 
                    f"{self.comp_lineEdits[f'comp_lineEdit_{i+1}'].text()}\t"+\
                    f"{self.label_MWs[f'label_MW_{i+1}'].text()}\t"+\
                    f"{self.C_lineEdits[f'C_lineEdit_{i+1}'].text()}\t"+\
                    f"{self.C_combos[f'C_combo_{i+1}'].currentText()}\t"+\
                    f"{self.V_lineEdits[f'V_lineEdit_{i+1}'].text()}\t"+\
                    f"{self.V_combos[f'V_combo_{i+1}'].currentText()}\t"+\
                    f"{self.mass_lbs[f'mass_lb_{i+1}'].text()}\n")
                        
            self.statusbar.showMessage('Find Weight data saved as .txt', 2000)
        
        elif tab == "tab_mzPeakShape":
            self.statusbar.showMessage('Saving peak massif data as .txt')
            
            with open(filename[0], 'w') as saveFile:
                
                reworked_name = self.name_lineEdit.text()
                if reworked_name == '':
                    reworked_name = 'User Input'
                reworked_sequence = self.textEditSequence_mz.toPlainText().strip()
                reworked_sequence = reworked_sequence.replace('\n','')
                reworked_sequence = reworked_sequence.upper()
                
                saveFile.write("name\tsequence\tcharge\tadduct\tmodification\n"
                            f"{reworked_name}\t"+\
                            f"{reworked_sequence}\t"+\
                            f"{self.charge_sb.value()}\t"+\
                            f"{self.adduct_cb.currentText()}\t"+\
                            f"{self.mod_cb.currentText()}\n"+\
                            "m/z\tRelative_Abundance\n")
                
                for mz, relAbn in (self.peakDataset):
                    saveFile.write(f"{mz}\t{relAbn}\n")
            
            self.statusbar.showMessage('Peak massif data saved as .txt', 2000)

    def writeXLSX(self, filename, tab):
        if tab == "tab_FindWeight":
            self.statusbar.showMessage('Saving Find Weight data as .xlsx')
            
            wb = openpyxl.Workbook()
            sheet = wb.active
            headerA = sheet['A1']
            headerA.value = "Name"
            headerB = sheet['B1']
            headerB.value = "MW"
            headerC = sheet['C1']
            headerC.value = "Concentration"
            headerD = sheet['D1']
            headerD.value = "Concentration unit"
            headerE = sheet['E1']
            headerE.value = "Volume"
            headerF = sheet['F1']
            headerF.value = "Volume unit"
            headerG = sheet['G1']
            headerG.value = "mass"
            headerH = sheet['H1']
            headerH.value = "mass unit"
            
            for i in range(len(self.comp_lineEdits)):
                cellA = sheet.cell(row = i+2, column=1)
                cellA.value = self.comp_lineEdits[f'comp_lineEdit_{i+1}'].text()
                cellB = sheet.cell(row = i+2, column=2)
                try:cellB.value = float(self.label_MWs[f'label_MW_{i+1}'].text())
                except:cellB.value = self.label_MWs[f'label_MW_{i+1}'].text()
                cellC = sheet.cell(row = i+2, column = 3)
                try:cellC.value = float(self.C_lineEdits[f'C_lineEdit_{i+1}'].text())
                except:cellC.value = self.C_lineEdits[f'C_lineEdit_{i+1}'].text()
                cellD = sheet.cell(row = i+2, column = 4)
                cellD.value = self.C_combos[f'C_combo_{i+1}'].currentText()
                cellE = sheet.cell(row = i+2, column=5)
                try:cellE.value = float(self.V_lineEdits[f'V_lineEdit_{i+1}'].text())
                except:cellE.value = self.V_lineEdits[f'V_lineEdit_{i+1}'].text()
                cellF = sheet.cell(row = i+2, column = 6)
                cellF.value = self.V_combos[f'V_combo_{i+1}'].currentText()
                cellG = sheet.cell(row = i+2, column = 7)
                try:cellG.value = float(self.mass_lbs[f'mass_lb_{i+1}'].text().split(' ')[0])
                except:cellG.value = self.mass_lbs[f'mass_lb_{i+1}'].text().split(' ')[0]
                cellH = sheet.cell(row = i+2, column = 8)
                try:cellH.value = self.mass_lbs[f'mass_lb_{i+1}'].text().split(' ')[1]
                except:pass
                       
            wb.save(filename[0])
            self.statusbar.showMessage('Find Weight data saved as .xlsx', 2000) 
            
        elif tab == "tab_mzPeakShape":
            self.statusbar.showMessage('Saving peak massif data as .xlsx')
            wb = openpyxl.Workbook()
            sheet = wb.active
            parameters_headerA = sheet['A1']
            parameters_headerA.value = "name"
            parameters_headerB = sheet['B1']
            parameters_headerB.value = "sequence"
            parameters_headerC = sheet['C1']
            parameters_headerC.value = "charge"
            parameters_headerD = sheet['D1']
            parameters_headerD.value = "adduct"
            parameters_headerE = sheet['E1']
            parameters_headerE.value = "modification"
            parametersA = sheet['A2'] 
            if self.name_lineEdit.text() == '':parametersA.value = 'User Input'
            else:parametersA.value = self.name_lineEdit.text()
            parametersB = sheet['B2']
            reworked_sequence = self.textEditSequence_mz.toPlainText().strip()
            reworked_sequence = reworked_sequence.replace('\n','')
            reworked_sequence = reworked_sequence.upper()
            parametersB.value = reworked_sequence
            parametersC = sheet['C2']
            parametersC.value = self.charge_sb.value()
            parametersD = sheet['D2']
            parametersD.value = self.adduct_cb.currentText()
            parametersE = sheet['E2']
            parametersE.value = self.mod_cb.currentText()            
            
            headerA = sheet['A3']
            headerA.value = "m/z"
            headerB = sheet['B3']
            headerB.value = "Relative Abundance"
            for index, data in enumerate(self.peakDataset):
                cellA = sheet.cell(row = index+4, column=1)
                cellA.value = data[0]
                cellB = sheet.cell(row = index+4, column =2)
                cellB.value = data[1]
            wb.save(filename[0])
            self.statusbar.showMessage('Peak massif data saved as .xlsx', 2000)
        else:
            self.statusbar.showMessage('Still under construction :)', 2000)
            
        
    def printPDF(self, filename, widget):
        self.statusbar.showMessage('Saving data as .pdf')
        
        printer = QPrinter(QPrinter.HighResolution)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(filename[0])
        painter = QtGui.QPainter(printer)
        
        # start scale
        xscale = printer.pageRect().width() * 0.9 / widget.width()
        yscale = printer.pageRect().height() * 0.9 / widget.height()
        scale = min(xscale, yscale)
        painter.translate(printer.paperRect().center())
        painter.scale(scale, scale)
        painter.translate(-widget.width() / 2, -widget.height() / 2)
        # end scale
        
        widget.render(painter)
        painter.end()
        
        self.statusbar.showMessage('Find Weight data saved as .pdf', 2000)
    
    def writeFASTA(self, filename):
        self.statusbar.showMessage('Saving FASTA sequence ...')
        with open(filename[0], 'w') as saveFile:
            saveFile.write(self.Uniprot_output_lb.text())
        self.statusbar.showMessage('FASTA sequence saved', 2000)

    def writeTextInfo(self, filename, tab):
        if tab == "tab_FindProt":
            self.statusbar.showMessage('Saving UniProt informations as .txt')
            with open(filename[0], 'w') as saveFile:
                saveFile.write(self.UP_ui.UniProtTextLabel.text())
            self.statusbar.showMessage('UniProt information saved', 2000)
            
        elif tab == "tab_inSilicoDigestion":
            self.statusbar.showMessage('Saving digestion output as .fasta')
            with open(filename[0], 'w') as saveFile:
                saveFile.write(self.UP_ui.UniProtTextLabel.text())
            self.statusbar.showMessage('digestion output saved', 2000)
            
    def fileNotRecognized(self):
        warningbox4 = QtWidgets.QMessageBox()
        warningbox4.setIcon(QtWidgets.QMessageBox.Warning)
        warningbox4.setWindowIcon(self.mainIcon)
        warningbox4.setWindowTitle('Error: file not recognized')
        warningbox4.setText("the file you are trying to open is not recognized.")
        warningbox4.exec_()
        self.statusbar.showMessage('opening file operation aborted', 2000)
        delattr(self, 'openFileName')

    def openFile(self, tab):
        if tab == "tab_FindWeight":
            self.openFileName = QtWidgets.QFileDialog().getOpenFileName(
                MainWindow, 
                'Open file',
                '',
                'text file (*.txt);; Excel sheet (*.xlsx)')
            if self.openFileName[1] == "text file (*.txt)":
                with open(self.openFileName[0], 'r') as openFile:
                    lines = openFile.read().splitlines()
                    header = lines.pop(0)
                    self.statusbar.showMessage('Opening file...')
                    if header == "Name\tMW\tConcentration\tConcentration unit\tVolume\tVolume unit\tmass":
                        self.openFindWeightForm(len(lines))
                        for i, line in enumerate(lines):
                            items = line.split('\t')
                            self.comp_lineEdits[f'comp_lineEdit_{i+1}'].setText(items[0])
                            self.label_MWs[f'label_MW_{i+1}'].setText(items[1])
                            self.C_lineEdits[f'C_lineEdit_{i+1}'].setText(items[2])
                            self.C_combos[f'C_combo_{i+1}'].setCurrentText(items[3])
                            self.V_lineEdits[f'V_lineEdit_{i+1}'].setText(items[4])
                            self.V_combos[f'V_combo_{i+1}'].setCurrentText(items[5])
                            self.mass_lbs[f'mass_lb_{i+1}'].setText(items[6])
                        self.statusbar.showMessage('file successfully opened', 2000)
                            
                    else:self.fileNotRecognized()
                        
            elif self.openFileName[1] == "Excel sheet (*.xlsx)":
                loaded_wb = openpyxl.load_workbook(self.openFileName[0])
                loaded_sheet = loaded_wb.active
                row_nb = loaded_sheet.max_row
                self.statusbar.showMessage('Opening file...') 
                
                if loaded_sheet["A1"].value == "Name"\
                and loaded_sheet["B1"].value == "MW"\
                and loaded_sheet["C1"].value == "Concentration"\
                and loaded_sheet["D1"].value == "Concentration unit"\
                and loaded_sheet["E1"].value == "Volume"\
                and loaded_sheet["F1"].value == "Volume unit"\
                and loaded_sheet["G1"].value == "mass"\
                and loaded_sheet["H1"].value == "mass unit":
                    self.openFindWeightForm(row_nb-1)
                    for i in range(row_nb-1):
                        self.comp_lineEdits[f'comp_lineEdit_{i+1}'].setText(str(loaded_sheet[f"A{i+2}"].value))
                        self.label_MWs[f'label_MW_{i+1}'].setText(str(loaded_sheet[f"B{i+2}"].value))
                        self.C_lineEdits[f'C_lineEdit_{i+1}'].setText(str(loaded_sheet[f"C{i+2}"].value))
                        self.C_combos[f'C_combo_{i+1}'].setCurrentText(str(loaded_sheet[f"D{i+2}"].value))
                        self.V_lineEdits[f'V_lineEdit_{i+1}'].setText(str(loaded_sheet[f"E{i+2}"].value))
                        self.V_combos[f'V_combo_{i+1}'].setCurrentText(str(loaded_sheet[f"F{i+2}"].value))
                        self.mass_lbs[f'mass_lb_{i+1}'].setText(str(loaded_sheet[f"G{i+2}"].value) + " " + str(loaded_sheet[f"H{i+2}"].value))                    
                    self.statusbar.showMessage('file successfully opened', 2000)
                    
                else:self.fileNotRecognized()
                
        elif tab == "tab_FindProt": 
            self.openFileName = QtWidgets.QFileDialog().getOpenFileName(
                MainWindow, 
                'Open file',
                '',
                'text file (*.txt);; FASTA file (*.fasta)')
            if self.openFileName[0] == '':
                delattr(self, 'openFileName')
            else:
                if self.openFileName[1] == "text file (*.txt)":
                    with open(self.openFileName[0], 'r') as openFile:
                        info = openFile.read()
                    self.openTextWindow(info)
                    
                elif self.openFileName[1] == "FASTA file (*.fasta)":
                    with open(self.openFileName[0], 'r') as openFile:
                        fastaFile = openFile.read()
                    self.Uniprot_output_lb.setText(fastaFile)
                    self.Uniprot_output_lb.setFont(self.fastaFont)
                    self.Uniprot_output_lb.setWordWrap(True)
                    self.Uniprot_output_lb.setCursor(QtGui.QCursor(QtCore.Qt.IBeamCursor))
                    self.Uniprot_output_lb.adjustSize()
                    self.outputScrollAreaWidget.adjustSize()
          
                else:self.fileNotRecognized()
        elif tab == "tab_inSilicoDigestion":
            self.openFileName = QtWidgets.QFileDialog().getOpenFileName(
                MainWindow, 
                'Open file',
                '',
                'FASTA sequence (*.fasta)')
            if self.openFileName[0] == '':
                delattr(self, 'openFileName')
            else:
                with open(self.openFileName[0], 'r') as openFile:
                    info = openFile.read()
                self.openTextWindow(info)
            
        elif tab == "tab_mzPeakShape":
            self.openFileName = QtWidgets.QFileDialog().getOpenFileName(
                MainWindow, 
                'Open file',
                '',
                'text file (*.txt);; Excel sheet (*.xlsx)')
            if self.openFileName[1] == "text file (*.txt)":
                with open(self.openFileName[0], 'r') as openFile:
                    lines = openFile.read().split('\n')
                    parameters_header = lines.pop(0)
                    parameters = lines.pop(0)
                    self.statusbar.showMessage('Opening file...')
                    if parameters_header == "name\tsequence\tcharge\tadduct\tmodification":
                        parametersList = parameters.split('\t')
                        if parametersList[1] == "":
                            self.statusbar.showMessage("empty sequence, cannot open file")
                        else:
                            try:
                                self.openPeakShapeWindow(parametersList[0],
                                parametersList[1],int(parametersList[2]),
                                parametersList[3],parametersList[4])
                                self.statusbar.showMessage('file successfully opened', 2000)
                                self.name_lineEdit.setText(parametersList[0])
                                self.textEditSequence_mz.setText(parametersList[1])
                                self.charge_sb.setValue(int(parametersList[2]))
                                self.adduct_cb.setCurrentText(parametersList[3])
                                self.mod_cb.setCurrentText(parametersList[4])
                            except Exception as e:
                                self.statusbar.showMessage(f"Unexpected Error: {e}", 4000)
   
                    else:
                        self.fileNotRecognized()
                                        
            elif self.openFileName[1] == "Excel sheet (*.xlsx)":
                loaded_wb = openpyxl.load_workbook(self.openFileName[0])
                loaded_sheet = loaded_wb.active
                row_nb = loaded_sheet.max_row
                self.statusbar.showMessage('Opening file...') 
                
                if loaded_sheet["A1"].value == "name"\
                and loaded_sheet["B1"].value == "sequence"\
                and loaded_sheet["C1"].value == "charge"\
                and loaded_sheet["D1"].value == "adduct"\
                and loaded_sheet["E1"].value == "modification":
                
                    try:
                        self.openPeakShapeWindow(loaded_sheet["A2"].value,
                        loaded_sheet["B2"].value, loaded_sheet["C2"].value,
                        loaded_sheet["D2"].value, loaded_sheet["E2"].value,)
                        self.statusbar.showMessage('file successfully opened', 2000)
                        self.name_lineEdit.setText(loaded_sheet["A2"].value)
                        self.textEditSequence_mz.setText(loaded_sheet["B2"].value)
                        self.charge_sb.setValue(loaded_sheet["C2"].value)
                        self.adduct_cb.setCurrentText(loaded_sheet["D2"].value)
                        self.mod_cb.setCurrentText(loaded_sheet["E2"].value)                     
                    except Exception as e:
                        self.statusbar.showMessage(f"Unexpected Error: {e}", 4000)

                else:self.fileNotRecognized()
        
        else:self.statusbar.showMessage('still under construction :)', 2000)
            
    def noDataError(self):
        warningbox3 = QtWidgets.QMessageBox()
        warningbox3.setIcon(QtWidgets.QMessageBox.Warning)
        warningbox3.setWindowIcon(self.mainIcon)
        warningbox3.setWindowTitle('Error: no data stored')
        warningbox3.setText("No data are currently stored.\nThere is nothing to save.")
        warningbox3.exec_()

    def save(self, tab):
        if tab == "tab_FindWeight":
            if hasattr(self, "ui"):
                if hasattr(self, 'saveFileName'):
                    if self.saveFileName[1] == "text (*.txt)":
                        self.writeTXT(self.saveFileName)                       
                    elif self.saveFileName[1] == "PDF (*.pdf)":
                        self.printPDF(self.saveFileName, self.ui.gridLayoutWidget)
                    elif self.saveFileName[1] == "Excel sheet (*.xlsx)":
                        self.writeXLSX(self.saveFileName, tab)
                    else:
                        self.statusbar.showMessage("format not recognized, can't save", 2000)

                else:
                    self.saveAs(tab)
            else: self.noDataError()
        
        elif tab == "tab_FindProt":
            if self.fasta_rb.isChecked():
                if self.Uniprot_output_lb.text() != "":
                    if hasattr(self, 'saveFileNameFASTA'):
                        if self.saveFileNameFASTA[1] == "FASTA sequence (*fasta)":
                            self.writeFASTA(self.saveFileNameFASTA)
                        else:
                            self.statusbar.showMessage("format not recognized, can't save", 2000)
                    else:
                        self.saveAs(tab)
                else: self.noDataError()
            else:
                if hasattr(self, 'textWindow'):
                    if hasattr(self, 'saveFileNameUI'):
                        if self.saveFileNameUI[1] == "text (*.txt)":
                            self.writeTextInfo(self.saveFileNameUI, self.tabWidget)
                        else: 
                            self.statusbar.showMessage("format not recognized, can't save", 2000)
                    else:self.saveAs(tab)
                else: self.noDataError()  
                
        elif tab == "tab_inSilicoDigestion":
            if hasattr(self, 'digestionPlainText'):
                if hasattr(self, 'saveFileNameDig'):
                    if self.saveFileNameDig[1] == "FASTA sequence (*.fasta)":
                        self.writeTextInfo(self.saveFileNameDig, self.tabWidget)
                    else:
                        self.statusbar.showMessage("format not recognized, can't save", 2000)
                    
                else:self.saveAs(tab)
            else:self.noDataError()
            
        elif tab == "tab_mzPeakShape":
            if hasattr(self, "PeakShape_ui"):
                if hasattr(self, 'saveFileNamePeak'):
                    if self.saveFileNamePeak[1] == "text (*.txt)":
                        self.writeTXT(self.saveFileNamePeak, tab)                       
                    elif self.saveFileNamePeak[1] == "PDF (*.pdf)":
                        self.printPDF(self.saveFileNamePeak, self.PeakShape_ui.graph)
                    elif self.saveFileNamePeak[1] == "Excel sheet (*.xlsx)":
                        self.writeXLSX(self.saveFileNamePeak, tab)
                    else:
                        self.statusbar.showMessage("format not recognized, can't save", 2000)

                else:
                    self.saveAs(tab)
            else: self.noDataError()
        else:
            self.statusbar.showMessage('still under construction :)', 2000)
           
    def saveAs(self, tab):

        if tab == "tab_FindWeight":
            if hasattr(self, 'window'):
                self.saveFileName = QtWidgets.QFileDialog().getSaveFileName(
                    MainWindow,
                    'Save report of your search',
                    'untitled',
                    "text (*.txt);; PDF (*.pdf);; Excel sheet (*.xlsx)")
                if self.saveFileName[0] == '':
                    delattr(self, 'saveFileName')
                else:
                    if self.saveFileName[1] == "text (*.txt)":
                        self.writeTXT(self.saveFileName)
                    elif self.saveFileName[1] == "PDF (*.pdf)":
                        self.printPDF(self.saveFileName, self.window)
                    elif self.saveFileName[1] == "Excel sheet (*.xlsx)":
                        self.writeXLSX(self.saveFileName, tab)
                    else:
                        self.statusbar.showMessage("format not recognized, can't save", 2000)                              
            else: self.noDataError()
        
        elif tab == "tab_FindProt":
            if self.fasta_rb.isChecked():
                if self.Uniprot_output_lb.text() != "":
                    self.saveFileNameFASTA = QtWidgets.QFileDialog().getSaveFileName(
                        MainWindow,
                        'Save FASTA sequence(s)',
                        'untitled',
                        "FASTA sequence (*.fasta)")
                    if self.saveFileNameFASTA[0] == '':
                        delattr(self, 'saveFileNameFASTA')
                    else:
                        self.writeFASTA(self.saveFileNameFASTA)
                else: self.noDataError()
            else:
                if hasattr(self, 'textWindow'):
                    self.saveFileNameUI = QtWidgets.QFileDialog().getSaveFileName(
                        MainWindow,
                        'Save information from UniProt',
                        'untitled',
                        "text (*.txt)")
                    if self.saveFileNameUI[0] == '':
                        delattr(self, 'saveFileNameUI')
                    else:
                        self.writeTextInfo(self.saveFileNameUI, tab)
                else: self.noDataError()
        
        elif tab == "tab_inSilicoDigestion":
            if hasattr(self, 'digestionPlainText'):
                self.saveFileNameDig = QtWidgets.QFileDialog().getSaveFileName(
                    MainWindow, 
                    'Save FASTA sequence(s)',
                    'untitled',
                    'FASTA sequence (*.fasta)')
                if self.saveFileNameDig[0] == '':
                    delattr(self, 'saveFileNameDig')
                else:
                    self.writeTextInfo(self.saveFileNameDig, tab)    
            else:self.noDataError()
        
        elif tab == "tab_mzPeakShape":
            if hasattr(self, "PeakShape_ui"):
                self.saveFileNamePeak = QtWidgets.QFileDialog().getSaveFileName(
                    MainWindow,
                    'Save graphics',
                    'untitled',
                    'text (*.txt);; Excel (*.xlsx);; PDF file (*.pdf)')
                if self.saveFileNamePeak[0] == '':
                    delattr(self, 'saveFileNamePeak')
                else:
                    if self.saveFileNamePeak[1] == "text (*.txt)":
                        self.writeTXT(self.saveFileNamePeak, tab)
                    elif self.saveFileNamePeak[1] == "Excel (*.xlsx)":
                        self.writeXLSX(self.saveFileNamePeak, tab)
                    elif self.saveFileNamePeak[1] == "PDF file (*.pdf)":
                        self.printPDF(self.saveFileNamePeak, self.PeakShape_ui.graph)
                    else:
                        self.statusbar.showMessage("format not recognized, can't save", 2000)
            else:self.noDataError()
        else:
            self.statusbar.showMessage('still under construction :)', 2000)
    
    def openManual(self):
        sp.Popen(["notepad.exe", "data\\README.txt"])
    
    def switchColor(self, color):
        if color == 'light':
            MainWindow.setStyleSheet(qdarktheme.load_stylesheet("light"))
            self.lightMode = True
            self.statusbar.showMessage("switched to light mode", 2000)
            try:self.window.setStyleSheet(qdarktheme.load_stylesheet('light'))
            except:pass
            try:self.MW_window.setStyleSheet(qdarktheme.load_stylesheet('light'))
            except:pass
            try:self.textWindow.setStyleSheet(qdarktheme.load_stylesheet('light'))
            except:pass
            try:self.peakShapeWindow.setStyleSheet(qdarktheme.load_stylesheet('light'))
            except:pass
            
        elif color == 'dark':
            MainWindow.setStyleSheet(qdarktheme.load_stylesheet())
            self.lightMode = False
            try:self.window.setStyleSheet(qdarktheme.load_stylesheet('dark'))
            except:pass
            try:self.MW_window.setStyleSheet(qdarktheme.load_stylesheet('dark'))
            except:pass
            try:self.textWindow.setStyleSheet(qdarktheme.load_stylesheet('dark'))
            except:pass
            try:self.peakShapeWindow.setStyleSheet(qdarktheme.load_stylesheet('dark'))
            except:pass
            self.statusbar.showMessage("switched to dark mode", 2000)
        self.settings.setValue('LightMode', self.lightMode)
    
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())